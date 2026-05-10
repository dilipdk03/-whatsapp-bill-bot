"""
WhatsApp Bill → Excel Bot (Multi-image support)
------------------------------------------------
Flow:
  1. User sends multiple bill/receipt images on WhatsApp
  2. User sends "done" when finished
  3. Bot processes all images with Groq Vision API
  4. All data combined into ONE Excel sheet
  5. Excel file sent back via WhatsApp
"""

import os
import uuid
import json
import base64
import threading
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, send_from_directory
from twilio.rest import Client
from twilio.twiml.messaging_response import MessagingResponse
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

# ── Config ───────────────────────────────────────────────────────────────────
GROQ_API_KEY           = os.getenv("GROQ_API_KEY")
TWILIO_ACCOUNT_SID     = os.getenv("TWILIO_ACCOUNT_SID")
TWILIO_AUTH_TOKEN      = os.getenv("TWILIO_AUTH_TOKEN")
TWILIO_WHATSAPP_NUMBER = os.getenv("TWILIO_WHATSAPP_NUMBER")
PUBLIC_URL             = os.getenv("PUBLIC_URL")

FILES_DIR = "generated_files"
os.makedirs(FILES_DIR, exist_ok=True)

twilio_client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# ── In-memory session store ───────────────────────────────────────────────────
# { "whatsapp:+91xxxxxxxxxx": [ {url, type}, {url, type}, ... ] }
user_sessions = {}

# ── Prompt ───────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a bill and expense data extractor. Extract ALL data from the bill/receipt image and return ONLY a valid JSON object — no markdown, no backticks, no explanation whatsoever.

Return exactly this structure:
{
  "vendor": "store or restaurant name",
  "date": "date in YYYY-MM-DD format if possible",
  "bill_number": "invoice or receipt number or null",
  "category": "Food / Travel / Medical / Utilities / Shopping / etc.",
  "items": [
    { "description": "item name", "quantity": 1, "unit_price": 0.00, "total": 0.00 }
  ],
  "subtotal": 0.00,
  "tax": 0.00,
  "discount": 0.00,
  "tip": 0.00,
  "total": 0.00,
  "payment_method": "Cash / Card / UPI / etc. or null",
  "currency": "INR or USD etc.",
  "notes": "any other info or null"
}

Use null for missing fields. Return pure JSON only."""


# ── Step 1: Receive WhatsApp webhook ─────────────────────────────────────────
@app.route("/webhook", methods=["POST"])
def webhook():
    num_media   = int(request.form.get("NumMedia", 0))
    from_number = request.form.get("From")
    body        = request.form.get("Body", "").strip().lower()

    # ── User sends "done" → process all collected images
    if body == "done":
        if from_number not in user_sessions or len(user_sessions[from_number]) == 0:
            resp = MessagingResponse()
            resp.message("⚠️ No images found! Please send your bill photos first, then type *done*.")
            return str(resp)

        images = user_sessions.pop(from_number)
        count  = len(images)
        send_whatsapp_message(from_number, f"⏳ Got it! Processing {count} bill(s)... please wait.")

        thread = threading.Thread(target=process_all_bills, args=(from_number, images))
        thread.start()
        return "", 200

    # ── User sends "cancel" → clear session
    if body == "cancel":
        user_sessions.pop(from_number, None)
        resp = MessagingResponse()
        resp.message("🗑️ Cleared! Send new bill photos whenever you're ready.")
        return str(resp)

    # ── Image received → add to session
    if num_media > 0:
        media_url  = request.form.get("MediaUrl0")
        media_type = request.form.get("MediaContentType0", "image/jpeg")

        if from_number not in user_sessions:
            user_sessions[from_number] = []

        user_sessions[from_number].append({"url": media_url, "type": media_type})
        count = len(user_sessions[from_number])

        resp = MessagingResponse()
        resp.message(
            f"📸 Bill {count} received!\n\n"
            f"Send more bill photos or type *done* to generate Excel with all {count} bill(s).\n"
            f"Type *cancel* to start over."
        )
        return str(resp)

    # ── No image, no command → send instructions
    resp = MessagingResponse()
    resp.message(
        "👋 Hi! I'm your *Bill to Excel* bot.\n\n"
        "📸 Send me one or multiple bill photos\n"
        "✅ Type *done* when finished\n"
        "🗑️ Type *cancel* to start over\n\n"
        "I'll combine all bills into one Excel file! 📊"
    )
    return str(resp)


# ── Step 2: Process all collected bills ──────────────────────────────────────
def process_all_bills(to_number, images):
    all_bills = []
    failed    = 0

    for i, img in enumerate(images, start=1):
        print(f"[DEBUG] Processing bill {i}/{len(images)}")
        try:
            image_b64 = download_twilio_image(img["url"])
            if not image_b64:
                failed += 1
                continue

            bill_data = extract_bill_with_groq(image_b64, img["type"])
            if bill_data:
                all_bills.append(bill_data)
            else:
                failed += 1
        except Exception as e:
            print(f"[ERROR] Bill {i} failed: {e}")
            failed += 1

    if not all_bills:
        send_whatsapp_message(to_number, "❌ Could not read any bills. Please send clearer images.")
        return

    # Build combined Excel
    filename = f"bills_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(FILES_DIR, filename)
    create_combined_excel(all_bills, filepath)

    file_url = f"{PUBLIC_URL}/files/{filename}"

    # Build summary
    total_amount = sum(b.get("total") or 0 for b in all_bills)
    currency     = all_bills[0].get("currency", "")
    summary      = (
        f"✅ *Done! {len(all_bills)} bill(s) extracted*\n\n"
        + "\n".join(
            f"🧾 Bill {i+1}: {b.get('vendor', 'Unknown')} — {b.get('currency','')} {b.get('total', 'N/A')}"
            for i, b in enumerate(all_bills)
        )
        + f"\n\n💰 *Grand Total: {currency} {round(total_amount, 2)}*"
        + (f"\n⚠️ {failed} image(s) could not be read." if failed else "")
        + "\n\n📥 Sending your Excel file now..."
    )

    send_whatsapp_message(to_number, summary)
    send_whatsapp_media(to_number, file_url, filename)


# ── Step 3: Download image from Twilio ───────────────────────────────────────
def download_twilio_image(media_url):
    try:
        response = requests.get(
            media_url,
            auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
            timeout=15
        )
        response.raise_for_status()
        return base64.b64encode(response.content).decode("utf-8")
    except Exception as e:
        print(f"[ERROR] Image download failed: {e}")
        return None


# ── Step 4: Groq Vision API ──────────────────────────────────────────────────
def extract_bill_with_groq(image_b64, media_type):
    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "meta-llama/llama-4-scout-17b-16e-instruct",
                "max_tokens": 1024,
                "messages": [
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{media_type};base64,{image_b64}"}
                            },
                            {"type": "text", "text": "Extract all bill/expense data from this image as JSON."}
                        ]
                    }
                ]
            },
            timeout=30
        )
        response.raise_for_status()
        raw_text = response.json()["choices"][0]["message"]["content"]
        clean    = raw_text.replace("```json", "").replace("```", "").strip()
        return json.loads(clean)
    except Exception as e:
        print(f"[ERROR] Groq API failed: {e}")
        return None


# ── Step 5: Build Combined Excel ─────────────────────────────────────────────
def create_combined_excel(bills, filepath):
    wb = openpyxl.Workbook()

    GREEN_DARK = "1A5276"
    GREEN_MID  = "D5F5E3"
    GRAY_LIGHT = "F2F3F4"
    WHITE      = "FFFFFF"
    DARK_TEXT  = "1C2833"

    def hf(): return Font(bold=True, color="FFFFFF", size=11)
    def sf(): return Font(bold=True, color=DARK_TEXT, size=10)
    def lf(): return Font(color="5D6D7E", size=10)
    def vf(): return Font(color=DARK_TEXT, size=10)
    def tf(): return Font(bold=True, color=DARK_TEXT, size=11)
    def fill(c): return PatternFill("solid", fgColor=c)
    def border():
        s = Side(style="thin", color="D5D8DC")
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal="center", vertical="center")
    def left():   return Alignment(horizontal="left",   vertical="center")
    def right():  return Alignment(horizontal="right",  vertical="center")

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.merge_cells("A1:G1")
    ws_summary["A1"] = "All Bills — Summary Report"
    ws_summary["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws_summary["A1"].fill      = fill(GREEN_DARK)
    ws_summary["A1"].alignment = center()
    ws_summary.row_dimensions[1].height = 28

    headers    = ["#", "Vendor", "Date", "Category", "Payment", "Currency", "Total"]
    col_widths = [5,    25,       15,      15,          15,        10,         12]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_summary.cell(row=2, column=col_idx, value=h)
        cell.font      = hf()
        cell.fill      = fill(GREEN_DARK)
        cell.alignment = center()
        cell.border    = border()
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = w
    ws_summary.row_dimensions[2].height = 22

    grand_total = 0
    currency    = ""

    for i, bill in enumerate(bills, start=1):
        bg       = WHITE if i % 2 == 0 else GRAY_LIGHT
        row      = i + 2
        total    = bill.get("total") or 0
        currency = bill.get("currency", "")
        grand_total += total

        values = [i, bill.get("vendor",""), bill.get("date",""), bill.get("category",""),
                  bill.get("payment_method",""), currency, total]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=val)
            cell.font      = vf()
            cell.fill      = fill(bg)
            cell.border    = border()
            cell.alignment = right() if col_idx in [1, 6, 7] else left()

    gt_row = len(bills) + 3
    ws_summary.merge_cells(f"A{gt_row}:F{gt_row}")
    ws_summary[f"A{gt_row}"] = "GRAND TOTAL"
    ws_summary[f"A{gt_row}"].font      = tf()
    ws_summary[f"A{gt_row}"].fill      = fill(GREEN_DARK)
    ws_summary[f"A{gt_row}"].alignment = right()
    ws_summary[f"G{gt_row}"] = round(grand_total, 2)
    ws_summary[f"G{gt_row}"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws_summary[f"G{gt_row}"].fill      = fill(GREEN_DARK)
    ws_summary[f"G{gt_row}"].alignment = right()
    for col in ["A","B","C","D","E","F","G"]:
        ws_summary[f"{col}{gt_row}"].border = border()
    ws_summary.row_dimensions[gt_row].height = 24
    ws_summary.freeze_panes = "A3"

    # ── Sheet per bill ────────────────────────────────────────────────────────
    for i, bill in enumerate(bills, start=1):
        vendor     = bill.get("vendor") or f"Bill {i}"
        sheet_name = f"Bill{i}_{vendor[:20]}".replace("/","").replace("\\","").replace("*","").replace("?","").replace("[","").replace("]","").replace(":","")[:31]
        ws   = wb.create_sheet(title=sheet_name)
        curr = bill.get("currency", "")
        row  = 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"Bill {i} — {vendor}"
        ws[f"A{row}"].font      = Font(bold=True, color="FFFFFF", size=12)
        ws[f"A{row}"].fill      = fill(GREEN_DARK)
        ws[f"A{row}"].alignment = center()
        ws.row_dimensions[row].height = 25
        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "Bill Information"
        ws[f"A{row}"].font      = sf()
        ws[f"A{row}"].fill      = fill(GREEN_MID)
        ws[f"A{row}"].alignment = left()
        row += 1

        for label, value in [("Vendor", bill.get("vendor")), ("Date", bill.get("date")),
                              ("Bill Number", bill.get("bill_number")), ("Category", bill.get("category")),
                              ("Payment Method", bill.get("payment_method")), ("Currency", bill.get("currency"))]:
            if value is None:
                continue
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = lf()
            ws[f"A{row}"].alignment = left()
            ws.merge_cells(f"B{row}:F{row}")
            ws[f"B{row}"] = str(value)
            ws[f"B{row}"].font = vf()
            ws[f"B{row}"].alignment = left()
            bg = WHITE if row % 2 == 0 else GRAY_LIGHT
            for col in ["A","B","C","D","E","F"]:
                ws[f"{col}{row}"].fill   = fill(bg)
                ws[f"{col}{row}"].border = border()
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 30
            row += 1

        row += 1

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "Line Items"
        ws[f"A{row}"].font      = sf()
        ws[f"A{row}"].fill      = fill(GREEN_MID)
        ws[f"A{row}"].alignment = left()
        row += 1

        for col_idx, (h, w) in enumerate(zip(["#","Description","Quantity","Unit Price","Total",""], [5,30,12,14,14,5]), start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = hf(); cell.fill = fill(GREEN_DARK)
            cell.alignment = center(); cell.border = border()
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        for j, item in enumerate(bill.get("items") or [], start=1):
            bg = WHITE if j % 2 == 0 else GRAY_LIGHT
            for col_idx, val in enumerate([j, item.get("description",""), item.get("quantity",""),
                f"{curr} {item.get('unit_price','')}" if item.get("unit_price") is not None else "",
                f"{curr} {item.get('total','')}" if item.get("total") is not None else "", ""], start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = vf(); cell.fill = fill(bg); cell.border = border()
                cell.alignment = right() if col_idx >= 3 else left()
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "Summary"
        ws[f"A{row}"].font = sf(); ws[f"A{row}"].fill = fill(GREEN_MID); ws[f"A{row}"].alignment = left()
        row += 1

        for label, value in [("Subtotal", bill.get("subtotal")), ("Tax", bill.get("tax")),
                              ("Discount", bill.get("discount")), ("Tip", bill.get("tip"))]:
            if value is None: continue
            ws.merge_cells(f"A{row}:D{row}"); ws[f"A{row}"] = label
            ws[f"A{row}"].font = lf(); ws[f"A{row}"].alignment = right()
            ws.merge_cells(f"E{row}:F{row}"); ws[f"E{row}"] = f"{curr} {value}"
            ws[f"E{row}"].font = vf(); ws[f"E{row}"].alignment = right()
            bg = WHITE if row % 2 == 0 else GRAY_LIGHT
            for col in ["A","B","C","D","E","F"]:
                ws[f"{col}{row}"].fill = fill(bg); ws[f"{col}{row}"].border = border()
            row += 1

        ws.merge_cells(f"A{row}:D{row}"); ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = tf(); ws[f"A{row}"].fill = fill(GREEN_DARK); ws[f"A{row}"].alignment = right()
        ws.merge_cells(f"E{row}:F{row}"); ws[f"E{row}"] = f"{curr} {bill.get('total','')}"
        ws[f"E{row}"].font = Font(bold=True, color="FFFFFF", size=12)
        ws[f"E{row}"].fill = fill(GREEN_DARK); ws[f"E{row}"].alignment = right()
        for col in ["A","B","C","D","E","F"]: ws[f"{col}{row}"].border = border()
        ws.row_dimensions[row].height = 24
        ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"[INFO] Combined Excel saved: {filepath}")


# ── Messaging helpers ─────────────────────────────────────────────────────────
def send_whatsapp_message(to, body):
    twilio_client.messages.create(from_=TWILIO_WHATSAPP_NUMBER, to=to, body=body)

def send_whatsapp_media(to, media_url, filename):
    twilio_client.messages.create(
        from_=TWILIO_WHATSAPP_NUMBER, to=to,
        body=f"📎 *Your Excel file is ready!*\n\nDownload it here:\n{media_url}\n\nOpen in Excel or Google Sheets."
    )

# ── Serve files ───────────────────────────────────────────────────────────────
@app.route("/files/<filename>")
def serve_file(filename):
    return send_from_directory(FILES_DIR, filename, as_attachment=True)

@app.route("/", methods=["GET"])
def health():
    return {"status": "running", "bot": "WhatsApp Bill → Excel (Multi)"}, 200

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)